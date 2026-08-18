from __future__ import annotations

from copy import copy
from pathlib import Path
import re

from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parents[3]
REPORT5 = ROOT / "My Drive" / "Report5"

UNIT = REPORT5 / "Report5_Unit Test.xlsx"
TEST_REPORT = REPORT5 / "Report5_Test Report.xlsx"


def copy_cell_format(source, target) -> None:
    if source.has_style:
        target._style = copy(source._style)
    if source.number_format:
        target.number_format = source.number_format
    if source.alignment:
        target.alignment = copy(source.alignment)
    if source.protection:
        target.protection = copy(source.protection)
    if source.font:
        target.font = copy(source.font)
    if source.fill:
        target.fill = copy(source.fill)
    if source.border:
        target.border = copy(source.border)


def ensure_merge(ws, range_string: str) -> None:
    if range_string not in {str(rng) for rng in ws.merged_cells.ranges}:
        ws.merge_cells(range_string)


def repair_unit_f007() -> None:
    wb = load_workbook(UNIT)
    ws = wb["F007_CreateWorkspace"]
    style_ws = wb["F001_Login"]

    # F007 was shifted up by one row. Insert the missing metadata row only when
    # the broken signature is still present so the script is idempotent.
    if ws["A2"].value != "Function Code" and ws["A2"].value == "Created By":
        ws.insert_rows(2, 1)

    # Restore the template metadata area using the current workbook's correct
    # function-sheet style as the reference. This preserves visual language and
    # fixes content without rebuilding the worksheet.
    for row in range(2, 8):
        for col in range(1, 20):
            copy_cell_format(style_ws.cell(row=row, column=col), ws.cell(row=row, column=col))
        ws.row_dimensions[row].height = style_ws.row_dimensions[row].height

    for col in range(1, 20):
        letter = ws.cell(row=1, column=col).column_letter
        ws.column_dimensions[letter].width = style_ws.column_dimensions[letter].width

    for merge_range in (
        "A2:B2",
        "C2:D2",
        "E2:J2",
        "K2:S2",
        "A3:B3",
        "C3:D3",
        "E3:J3",
        "K3:S3",
        "A4:B4",
        "C4:D4",
        "E4:J4",
        "K4:S4",
        "A5:B5",
        "C5:S5",
        "A6:B6",
        "C6:D6",
        "E6:J6",
        "K6:M6",
        "N6:S6",
        "A7:B7",
        "C7:D7",
        "E7:J7",
        "K7:K7",
        "L7:L7",
        "M7:M7",
        "N7:S7",
    ):
        if ":" in merge_range and merge_range.split(":")[0] != merge_range.split(":")[1]:
            ensure_merge(ws, merge_range)

    ws["A2"] = "Function Code"
    ws["C2"] = "F007"
    ws["E2"] = "Function Name"
    ws["K2"] = "CreateWorkspaceAsync"

    ws["A3"] = "Created By"
    ws["C3"] = "Huynh Thai Tu"
    ws["E3"] = "Executed By"
    ws["K3"] = "Huynh Thai Tu"

    ws["A4"] = "Lines of code"
    ws["E4"] = "Lack of test cases"
    ws["K4"] = 0

    ws["A5"] = "Test requirement"
    ws["C5"] = "Create workspace"

    ws["A6"] = "Passed"
    ws["C6"] = "Failed"
    ws["E6"] = "Untested"
    ws["K6"] = "N/A/B"
    ws["N6"] = "Total Test Cases"

    ws["A7"] = 7
    ws["C7"] = 0
    ws["E7"] = 0
    ws["K7"] = 6
    ws["L7"] = 1
    ws["M7"] = 0
    ws["N7"] = 7

    ws.freeze_panes = "E11"

    # Keep summary statistics aligned with the fixed function sheet.
    stats = wb["Statistics"]
    for row in range(9, 60):
        if stats.cell(row=row, column=2).value == "F007":
            stats.cell(row=row, column=3).value = "=F007_CreateWorkspace!A7"
            stats.cell(row=row, column=4).value = "=F007_CreateWorkspace!C7"
            stats.cell(row=row, column=5).value = "=F007_CreateWorkspace!E7"
            stats.cell(row=row, column=6).value = "=F007_CreateWorkspace!K7"
            stats.cell(row=row, column=7).value = "=F007_CreateWorkspace!L7"
            stats.cell(row=row, column=8).value = "=F007_CreateWorkspace!M7"
            stats.cell(row=row, column=9).value = f"=SUM(C{row}:E{row})"
            stats.cell(row=row, column=10).value = "Workspace Management"
            break

    wb.save(UNIT)


def repair_test_report() -> None:
    wb = load_workbook(TEST_REPORT)
    feature_sheets = [
        name
        for name in wb.sheetnames
        if name not in {"Cover", "Test Cases", "Test Statistics", "Traceability Matrix"}
    ]

    stats = wb["Test Statistics"]
    for index, sheet_name in enumerate(feature_sheets, start=13):
        stats.cell(row=index, column=3).value = sheet_name

    # Preserve the template layout, but make Round 2 and Round 3 status formulas
    # count their own result columns instead of inheriting Round 1 column F.
    for sheet_name in feature_sheets:
        ws = wb[sheet_name]
        for round_row, result_col in ((7, "I"), (8, "L")):
            base_formula = ws[f"B{round_row}"].value
            end_row = "999"
            if isinstance(base_formula, str):
                match = re.search(rf"\${result_col}\$?10:\${result_col}\$?(\d+)", base_formula)
                if match:
                    end_row = match.group(1)
            ws[f"B{round_row}"] = f'=COUNTIF(${result_col}$10:${result_col}${end_row},B5)'
            ws[f"C{round_row}"] = f'=COUNTIF(${result_col}$10:${result_col}${end_row},C5)'
            ws[f"D{round_row}"] = f'=COUNTIF(${result_col}$10:${result_col}${end_row},D5)'
            ws[f"E{round_row}"] = f'=COUNTIF(${result_col}$10:${result_col}${end_row},E5)'

    wb.save(TEST_REPORT)


def main() -> None:
    repair_unit_f007()
    repair_test_report()


if __name__ == "__main__":
    main()
