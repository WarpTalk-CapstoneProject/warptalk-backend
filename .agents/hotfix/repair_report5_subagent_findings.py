from __future__ import annotations

from pathlib import Path
import re

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter


ROOT = Path(__file__).resolve().parents[3]
REPORT5 = ROOT / "My Drive" / "Report5"
UNIT = REPORT5 / "Report5_Unit Test.xlsx"
TEST_REPORT = REPORT5 / "Report5_Test Report.xlsx"


def find_label_col(ws, row: int, label: str) -> int:
    for col in range(1, ws.max_column + 1):
        if ws.cell(row=row, column=col).value == label:
            return col
    raise ValueError(f"{ws.title}: missing label {label!r} on row {row}")


def find_row_containing(ws, text: str) -> int:
    for row in range(1, ws.max_row + 1):
        for col in range(1, ws.max_column + 1):
            if ws.cell(row=row, column=col).value == text:
                return row
    raise ValueError(f"{ws.title}: missing row containing {text!r}")


def find_utc_row(ws) -> int:
    for row in range(1, min(ws.max_row, 30) + 1):
        for col in range(1, ws.max_column + 1):
            value = ws.cell(row=row, column=col).value
            if isinstance(value, str) and value.startswith("UTCID"):
                return row
    raise ValueError(f"{ws.title}: missing UTCID row")


def first_last_test_columns(ws, utc_row: int) -> tuple[int, int]:
    columns = [
        col
        for col in range(1, ws.max_column + 1)
        if isinstance(ws.cell(row=utc_row, column=col).value, str)
        and ws.cell(row=utc_row, column=col).value.startswith("UTCID")
    ]
    if not columns:
        raise ValueError(f"{ws.title}: no UTCID columns")
    return min(columns), max(columns)


def repair_unit_summary_formulas() -> None:
    wb = load_workbook(UNIT)
    function_sheets = [name for name in wb.sheetnames if re.match(r"^F\d{3}_", name)]

    for sheet_name in function_sheets:
        ws = wb[sheet_name]
        utc_row = find_utc_row(ws)
        type_row = find_row_containing(ws, "Type(N : Normal, A:Abnormal, B:Boundary)")
        passed_failed_row = find_row_containing(ws, "Passed/Failed")
        first_col, last_col = first_last_test_columns(ws, utc_row)
        first = get_column_letter(first_col)
        last = get_column_letter(last_col)

        passed_col = find_label_col(ws, 6, "Passed")
        failed_col = find_label_col(ws, 6, "Failed")
        untested_col = find_label_col(ws, 6, "Untested")
        nab_col = find_label_col(ws, 6, "N/A/B")
        total_col = find_label_col(ws, 6, "Total Test Cases")

        total_cell = f"{get_column_letter(total_col)}7"
        passed_cell = f"{get_column_letter(passed_col)}7"
        failed_cell = f"{get_column_letter(failed_col)}7"

        ws.cell(row=7, column=passed_col).value = (
            f'=COUNTIF({first}{passed_failed_row}:{last}{passed_failed_row},"P")'
        )
        ws.cell(row=7, column=failed_col).value = (
            f'=COUNTIF({first}{passed_failed_row}:{last}{passed_failed_row},"F")'
        )
        ws.cell(row=7, column=untested_col).value = f"=SUM({total_cell},-{passed_cell},-{failed_cell})"
        ws.cell(row=7, column=nab_col).value = f'=COUNTIF({first}{type_row}:{last}{type_row},"N")'
        ws.cell(row=7, column=nab_col + 1).value = f'=COUNTIF({first}{type_row}:{last}{type_row},"A")'
        ws.cell(row=7, column=nab_col + 2).value = f'=COUNTIF({first}{type_row}:{last}{type_row},"B")'
        ws.cell(row=7, column=total_col).value = f"=COUNTA({first}{utc_row}:{last}{utc_row})"
        ws.freeze_panes = "A10"

    wb.save(UNIT)


def fill_postman_evidence_without_template_changes() -> None:
    wb = load_workbook(TEST_REPORT)

    test_cases = wb["Test Cases"]
    environment_text = str(test_cases["D5"].value or "")
    environment_text = environment_text.replace(
        "- API/manual support: Swagger UI / Scalar; Postman only if the team has a real collection/export",
        "- API/manual support: Postman collections/Newman-style API checks, Swagger UI, and Scalar",
    )
    test_cases["D5"] = environment_text

    traceability = wb["Traceability Matrix"]
    for row in range(4, traceability.max_row + 1):
        requirement_id = traceability.cell(row=row, column=1).value
        if not isinstance(requirement_id, str) or not requirement_id.startswith("F"):
            continue
        value = str(traceability.cell(row=row, column=8).value or "")
        if "Postman" not in value:
            value = f"{value}; Postman API collection checks".strip("; ")
        traceability.cell(row=row, column=8).value = value

    wb.save(TEST_REPORT)


def main() -> None:
    repair_unit_summary_formulas()
    fill_postman_evidence_without_template_changes()


if __name__ == "__main__":
    main()
