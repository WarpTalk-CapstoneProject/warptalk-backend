from __future__ import annotations

from pathlib import Path
import json

from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parents[3]
REPORT5 = ROOT / "My Drive" / "Report5"
TEST_REPORT = REPORT5 / "Report5_Test Report.xlsx"
POSTMAN_COLLECTION = ROOT / "warptalk-backend" / "postman" / "WarpTalk-Backend.postman_collection.json"
POSTMAN_ENV = (
    ROOT
    / "warptalk-backend"
    / "postman"
    / "environments"
    / "WarpTalk-Backend.Local.postman_environment.json"
)


def postman_summary() -> tuple[int, dict[str, int]]:
    data = json.loads(POSTMAN_COLLECTION.read_text(encoding="utf-8"))
    counts: dict[str, int] = {}

    def walk(items: list[dict], trail: tuple[str, ...] = ()) -> None:
        for item in items:
            name = item.get("name", "")
            if "item" in item:
                walk(item["item"], trail + (name,))
            else:
                top = trail[0] if trail else "(root)"
                counts[top] = counts.get(top, 0) + 1

    walk(data.get("item", []))
    return sum(counts.values()), counts


def upsert_trace_row(ws, requirement_id: str, values: list[str]) -> None:
    target_row = None
    for row in range(4, ws.max_row + 1):
        if ws.cell(row=row, column=1).value == requirement_id:
            target_row = row
            break
    if target_row is None:
        target_row = 4
        while ws.cell(row=target_row, column=1).value not in (None, ""):
            target_row += 1

    for col, value in enumerate(values, start=1):
        ws.cell(row=target_row, column=col).value = value


def main() -> None:
    request_count, module_counts = postman_summary()
    wb = load_workbook(TEST_REPORT)

    test_cases = wb["Test Cases"]
    environment_text = str(test_cases["D5"].value or "")
    environment_text = environment_text.replace(
        "- API/manual support: Postman collections/Newman-style API checks, Swagger UI, and Scalar",
        (
            "- API/manual support: Postman collection "
            "`warptalk-backend/postman/WarpTalk-Backend.postman_collection.json` "
            f"({request_count} requests) with local environment "
            "`warptalk-backend/postman/environments/WarpTalk-Backend.Local.postman_environment.json`; "
            "Swagger UI and Scalar for endpoint inspection. No Newman run-log claim is made in this report."
        ),
    )
    test_cases["D5"] = environment_text

    traceability = wb["Traceability Matrix"]
    postman_tool = (
        f"Postman collection ({request_count} requests) at "
        "warptalk-backend/postman/WarpTalk-Backend.postman_collection.json; "
        "local environment at warptalk-backend/postman/environments/WarpTalk-Backend.Local.postman_environment.json"
    )

    for row in range(4, traceability.max_row + 1):
        requirement_id = traceability.cell(row=row, column=1).value
        if not isinstance(requirement_id, str) or not requirement_id.startswith("F"):
            continue
        value = str(traceability.cell(row=row, column=8).value or "")
        value = value.replace("Postman API collection checks", postman_tool)
        if "Postman" not in value:
            value = f"{value}; {postman_tool}".strip("; ")
        traceability.cell(row=row, column=8).value = value

    uat_rows = [
        [
            "UAT-UC-01",
            "UAT",
            "Account onboarding and session recovery",
            "Authentication; Token Refresh; User Settings",
            "User can register/login, refresh session, update settings, and recover from auth failures.",
            "F001/F003/F004/F005/F033",
            "warptalk-backend/auth/tests; warptalk-backend/postman",
            f"{postman_tool}; Linear WT-344/WT-361/WT-405 bug-regression references in auth tests",
            "Derived from actual auth use cases and bug-regression tickets.",
        ],
        [
            "UAT-UC-02",
            "UAT",
            "Workspace creation, invitation, membership, and domain governance",
            "Workspace Management; Workspace Invitation; Workspace Member Role; Workspace Domain",
            "Owner/Admin creates workspace, invites member, accepts invitation, changes role, and validates domain rules.",
            "F006/F007/F008/F009/F010/F031/F035",
            "warptalk-backend/workspace/tests; warptalk-backend/postman",
            f"{postman_tool}; Linear WT-313/WT-349/WT-352/WT-371/WT-417 references",
            "UAT uses real workspace flows plus Linear bug-ticket regressions around role/domain/access behavior.",
        ],
        [
            "UAT-UC-03",
            "UAT",
            "Translation room lifecycle and participant access",
            "Room Participant; Translation Room Lifecycle; Translation Room Join; Meeting Lifecycle",
            "Host creates/starts/ends room, participant joins/admitted, waiting-room and owner/admin access behave correctly.",
            "F011/F012/F013/F015/F017/F018/F032",
            "warptalk-backend/translation-room/tests; warptalk-backend/meeting/tests; warptalk-backend/postman",
            f"{postman_tool}; Linear WT-304/WT-313/WT-371/WT-386 references",
            "UAT follows actual room/meeting flows and regression tickets for participant/host authority.",
        ],
        [
            "UAT-UC-04",
            "UAT",
            "Live transcript, translation, audio route, and voice clone consent",
            "Voice Clone Consent; Translation Room Lifecycle; Assistant Q&A",
            "User grants consent, starts translation, sees live transcript/translation continue, and validates degraded/retry behavior.",
            "F012/F014/F034",
            "warptalk-backend/translation-room/tests; warptalk-ai/tests; warptalk-backend/postman",
            f"{postman_tool}; Linear WT-371/WT-382/WT-387/WT-389 references",
            "UAT uses real live meeting use cases and bug tickets for stale session, transcript stop, and voice-output instability.",
        ],
        [
            "UAT-UC-05",
            "UAT",
            "Meeting collaboration: chat, poll, Q&A, breakout",
            "Messaging; Poll Management; Poll Voting; Question List; Session Start",
            "Participant sends chat/Q&A, votes in poll, host manages breakout/session actions.",
            "F016/F019/F020/F026/F030",
            "warptalk-backend/meeting/tests; warptalk-backend/postman",
            postman_tool,
            "UAT follows actual meeting collaboration use cases from system/API sheets.",
        ],
        [
            "UAT-UC-06",
            "UAT",
            "Billing lifecycle, credits, usage, webhook, and workspace suspension",
            "Billing Authorization; Credit Consumption; Webhook Handling; Usage Recording; Contract Terms; Workspace Suspension",
            "Workspace owner checks plan/entitlement, consumes credits, processes payment/webhook, and verifies suspension/retention behavior.",
            "F021/F022/F023/F024/F025/F031",
            "warptalk-backend/billing/tests; warptalk-backend/workspace/tests; warptalk-backend/postman",
            f"{postman_tool}; Linear WT-349/WT-381 references",
            "UAT references real billing use cases and Linear bug tickets about lifecycle, expiry, retention, and workspace state.",
        ],
        [
            "UAT-UC-07",
            "UAT",
            "Notification read state and navigation",
            "Notification Read Status; Validation Test; Validation",
            "User receives notification, opens target flow, marks as read, and admin validation gates hold.",
            "F027/F028/F029",
            "warptalk-backend/notification/tests; warptalk-backend/postman",
            f"{postman_tool}; Linear WT-261/WT-364 references",
            "UAT references actual notification use cases plus Linear bug-ticket navigation/read-state regressions.",
        ],
        [
            "UAT-UC-08",
            "UAT",
            "Knowledge, transcript, glossary, and assistant retrieval",
            "Workspace Knowledge; Assistant Q&A; Transcript-related system/API sheets",
            "User uploads/uses workspace knowledge, asks assistant, and validates transcript/glossary retrieval boundaries.",
            "F034/F035",
            "warptalk-backend/workspace/tests; warptalk-ai/tests; warptalk-backend/postman",
            f"{postman_tool}; Linear WT-240/WT-241/WT-371 references",
            "UAT takes actual knowledge/assistant use cases and prior testability/maintainability bug-ticket scope.",
        ],
    ]

    for row in uat_rows:
        upsert_trace_row(traceability, row[0], row)

    wb.save(TEST_REPORT)


if __name__ == "__main__":
    main()
